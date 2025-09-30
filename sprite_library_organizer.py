#!/usr/bin/env python3
"""
sprite_library_organizer.py

Organize finalized Student Transfer character folders into a library:
~/Documents/ST Sprite Library/<Boys|Girls|Unknown>/<game>/<display_name>

Improvements vs prior version:
- If <display_name> already exists in the destination bucket/game folder,
  select a NEW gendered name from names.xlsx (or names.csv) at random until unique.
- After copying, update the copied character.yml's display_name to match the final folder name.
- Non-binary/unknown voices go to "Unknown" bucket and use a neutral name pool if needed.

Expected names file schema:
- names.xlsx OR names.csv with columns: name, gender
  (gender can be girl/boy/female/male; case-insensitive; extra columns ignored)

Usage:
  python sprite_library_organizer.py /path/to/finalized/chars \
      --library-root "C:/Users/You/Documents/ST Sprite Library" \
      --names-file names.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
import random
import shutil
from pathlib import Path
from typing import Tuple, List, Dict, Optional

import yaml

DEF_LIBRARY_NAME = "ST Sprite Library"
DEF_NAMES_XLSX = "names.xlsx"
DEF_NAMES_CSV  = "names.csv"

# -----------------------------
# Basic FS helpers
# -----------------------------
from pathlib import Path
import os, re

def _documents_folder() -> Path:
    # 0) Optional explicit override
    env = os.environ.get("ST_SPRITE_LIBRARY_ROOT")
    if env:
        return Path(os.path.expandvars(os.path.expanduser(env)))

    # 1) Windows: use registry (handles localized/non-standard Documents path)
    if os.name == "nt":
        try:
            import winreg
            with winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders"
            ) as k:
                doc = winreg.QueryValueEx(k, "Personal")[0]
                if doc:
                    return Path(doc)
        except Exception:
            pass
        return Path.home() / "Documents"

    # 2) Linux: honor XDG user dirs if present
    try:
        xdg = Path.home() / ".config/user-dirs.dirs"
        if xdg.exists():
            txt = xdg.read_text(encoding="utf-8", errors="ignore")
            m = re.search(r'XDG_DOCUMENTS_DIR="?(.+?)"?', txt)
            if m:
                raw = m.group(1).replace("$HOME", str(Path.home()))
                return Path(os.path.expandvars(raw)).expanduser()
    except Exception:
        pass

    # 3) macOS & fallback
    return Path.home() / "Documents"


def _sanitize(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r"[\\/:\*\?\"<>\|]+", "_", name)  # Windows-invalid chars
    name = re.sub(r"\s+", " ", name).strip()
    return name or "untitled"

def _read_yaml(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except Exception:
        return {}

def _write_yaml(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        yaml.dump(data, f, sort_keys=False, allow_unicode=True)

def _gender_bucket(voice: Optional[str]) -> str:
    """Return Boys/Girls/Unknown for library bucketing."""
    v = (voice or "").strip().lower()
    # common aliases
    if v in {"male", "boy", "man", "m", "john", "kiyoshi"}:
        return "Boys"
    if v in {"female", "girl", "woman", "f", "tomboy", "johngb"}:
        return "Girls"
    return "Unknown"

def _voice_to_simple_gender(voice: Optional[str]) -> Optional[str]:
    """Return 'girl', 'boy', or None for name-pool selection."""
    v = (voice or "").strip().lower()
    if v in {"female", "girl", "woman", "f", "tomboy", "johngb"}:
        return "girl"
    if v in {"male", "boy", "man", "m", "john", "kiyoshi"}:
        return "boy"
    return None

# -----------------------------
# Name pool loading (xlsx/csv)
# -----------------------------
def _load_names_from_xlsx(xlsx_path: Path) -> Tuple[List[str], List[str]]:
    try:
        import openpyxl  # lazy import
    except Exception:
        print(f"[WARN] openpyxl not available; cannot read '{xlsx_path.name}'.")
        return [], []

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        # find header indexes
        headers = {str(c.value).strip().lower(): idx for idx, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1))[0:])}
        # fallback if openpyxl returns tuples differently
        if not headers:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = {str(v).strip().lower(): i for i, v in enumerate(header_row)}

        name_idx = headers.get("name")
        gender_idx = headers.get("gender")

        g_girl, g_boy = [], []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            name = str(row[name_idx]).strip() if (name_idx is not None and row[name_idx] is not None) else ""
            gender = str(row[gender_idx]).strip().lower() if (gender_idx is not None and row[gender_idx] is not None) else ""
            if not name:
                continue
            if gender in {"girl", "female", "f"}:
                g_girl.append(name)
            elif gender in {"boy", "male", "m"}:
                g_boy.append(name)
        return g_girl, g_boy
    except Exception as e:
        print(f"[WARN] Failed reading xlsx '{xlsx_path}': {e}")
        return [], []

def _load_names_from_csv(csv_path: Path) -> Tuple[List[str], List[str]]:
    import csv
    g_girl, g_boy = [], []
    try:
        with csv_path.open(newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = (row.get("name") or "").strip()
                gender = (row.get("gender") or "").strip().lower()
                if not name:
                    continue
                if gender in {"girl", "female", "f"}:
                    g_girl.append(name)
                elif gender in {"boy", "male", "m"}:
                    g_boy.append(name)
    except Exception as e:
        print(f"[WARN] Failed reading csv '{csv_path}': {e}")
    return g_girl, g_boy

def _load_name_pool(preferred: Optional[Path]) -> Tuple[List[str], List[str], List[str]]:
    """
    Returns (girl_names, boy_names, neutral_names).
    - Will read preferred if set, otherwise try names.xlsx -> names.csv in CWD.
    - Neutral fallbacks are only used for Unknown bucket.
    """
    girl, boy = [], []

    if preferred:
        p = preferred.expanduser().resolve()
        if p.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            girl, boy = _load_names_from_xlsx(p)
        elif p.suffix.lower() == ".csv":
            girl, boy = _load_names_from_csv(p)
        else:
            print(f"[WARN] Unsupported names file type: {p.suffix}")

    if not girl and not boy:
        here = Path.cwd()
        xlsx = here / DEF_NAMES_XLSX
        csvf = here / DEF_NAMES_CSV
        if xlsx.exists():
            girl, boy = _load_names_from_xlsx(xlsx)
        elif csvf.exists():
            girl, boy = _load_names_from_csv(csvf)

    if not girl and not boy:
        print("[WARN] No valid names file found. Using small hardcoded pool.")
        girl = ["Sakura", "Emily", "Yuki", "Hannah", "Aiko", "Madison", "Kana", "Sara"]
        boy  = ["Takashi", "Ethan", "Yuto", "Liam", "Kenta", "Jacob", "Hiro", "Alex"]

    neutral = ["Alex", "Riley", "Taylor", "Jordan", "Casey", "Morgan", "Sam", "Jamie"]
    return girl, boy, neutral

# -----------------------------
# Core copy + naming logic
# -----------------------------
def _choose_unique_name(dest_parent: Path, base_display_name: str,
                        gender_simple: Optional[str],
                        girl_pool: List[str], boy_pool: List[str], neutral_pool: List[str]) -> str:
    """
    If 'base_display_name' is free, return it.
    Otherwise, repeatedly draw names from the appropriate pool until unique.
    If the pool is exhausted / unlucky, fall back to numeric suffix.
    """
    desired = _sanitize(base_display_name)
    if not (dest_parent / desired).exists():
        return desired

    # pick pool
    if gender_simple == "girl":
        pool = list(dict.fromkeys(girl_pool))  # de-dup
    elif gender_simple == "boy":
        pool = list(dict.fromkeys(boy_pool))
    else:
        pool = list(dict.fromkeys(neutral_pool))

    random.shuffle(pool)
    tried = set([desired.lower()])

    for candidate in pool:
        cand = _sanitize(candidate)
        if cand.lower() in tried:
            continue
        if not (dest_parent / cand).exists():
            return cand
        tried.add(cand.lower())

    # fallback: numeric suffix on original
    i = 2
    while True:
        cand = f"{desired}_{i}"
        if not (dest_parent / cand).exists():
            return cand
        i += 1

def _copy_character(src_dir: Path, dest_dir: Path, *, dry_run: bool=False) -> None:
    print(f"[COPY] {src_dir.name} -> {dest_dir}")
    if dry_run:
        return
    shutil.copytree(src_dir, dest_dir)

def _update_yaml_display_name(dest_dir: Path, new_name: str, *, dry_run: bool=False) -> None:
    yml = dest_dir / "character.yml"
    meta = _read_yaml(yml)
    if not meta:
        return
    old = meta.get("display_name")
    if str(old) == str(new_name):
        return
    meta["display_name"] = new_name
    if dry_run:
        print(f"[DRY ] Would update {yml} display_name: '{old}' -> '{new_name}'")
        return
    try:
        _write_yaml(yml, meta)
        print(f"[INFO] Updated {yml.name} display_name: '{old}' -> '{new_name}'")
    except Exception as e:
        print(f"[WARN] Failed to update {yml}: {e}")

def organize_sprite_library(source_root: Path, library_root: Path,
                            names_file: Optional[Path] = None,
                            *, dry_run: bool=False) -> None:
    if not source_root.is_dir():
        raise SystemExit(f"[ERROR] '{source_root}' is not a directory.")

    girl_pool, boy_pool, neutral_pool = _load_name_pool(names_file)

    total = 0
    placed = 0

    for char_dir in sorted([p for p in source_root.iterdir() if p.is_dir()]):
        total += 1
        yml = char_dir / "character.yml"
        meta = _read_yaml(yml)

        # read metadata
        display_name = _sanitize(str(meta.get("display_name") or char_dir.name))
        voice = meta.get("voice")
        game = _sanitize(str(meta.get("game") or "_unknown_game"))

        bucket = _gender_bucket(voice)
        gender_simple = _voice_to_simple_gender(voice)

        dest_parent = library_root / bucket / game
        dest_parent.mkdir(parents=True, exist_ok=True)   # correct spot
        final_name = _choose_unique_name(
            dest_parent, display_name, gender_simple, girl_pool, boy_pool, neutral_pool
        )
        dest_dir = dest_parent / final_name


        # copy
        _copy_character(char_dir, dest_dir, dry_run=dry_run)
        placed += 1

        # ensure YAML in the copy reflects final folder name
        _update_yaml_display_name(dest_dir, final_name, dry_run=dry_run)

    print(f"\n[INFO] Finished. Characters scanned: {total}, placed: {placed}")
    print(f"[INFO] Library root: {library_root}")

# -----------------------------
# CLI
# -----------------------------
def main():
    ap = argparse.ArgumentParser(description="Organize character folders into the ST Sprite Library.")
    ap.add_argument("source_root", help="Path containing character folders to organize.")
    ap.add_argument("--library-root", default=None,
                    help="Destination root (default: ~/Documents/ST Sprite Library)")
    ap.add_argument("--names-file", default=None,
                    help="Path to names.xlsx or names.csv (default: try names.xlsx then names.csv in CWD)")
    ap.add_argument("--dry-run", action="store_true", help="Print actions without copying/writing.")
    args = ap.parse_args()

    source_root = Path(args.source_root).expanduser().resolve()
    library_root = Path(args.library_root).expanduser().resolve() if args.library_root else (_documents_folder() / DEF_LIBRARY_NAME)
    names_file = Path(args.names_file).expanduser().resolve() if args.names_file else None

    organize_sprite_library(source_root, library_root, names_file, dry_run=bool(args.dry_run))

if __name__ == "__main__":
    main()
